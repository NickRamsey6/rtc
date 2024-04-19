import pandas as pd
from openpyxl import Workbook, load_workbook
from openpyxl.utils import get_column_letter

# Load workbook
wb = load_workbook('path')
ws = wb.active

# Access specific cells
print(ws['A1'].value)

# Modify specific cells
ws['A1'] = 'Test'
wb.save(path)

# Access specific sheets
print(wb.sheetnames)
print(wb['Sheet1'])
ws = wb['Sheet1']

# Create new sheet
wb.create_sheet('New_Sheet')
wb.save(path)

# Create new workbook
wb = Workbook()
ws = wb.active
ws.title = 'Data'
ws.append(['Col1', 'Col2', 'Col3'])
wb.save(path)

# Loop through rows and cols (10 rows and 5 columns in this example)
for row in range(1, 11):
    for col in range(1, 5):
        char = get_column_letter(col)
        print(ws[char + str(row)].value)
        # Change values need to save after
        # ws[char + str(row)] = 'new value'

# Merge cells
ws.merge_cells('A1:D1')
# ws.unmerge_cells('A1:D1')
wb.save(path)

# Insert Empty row
ws.insert_rows(7)
# Insert empty row after row 7

# Create new csv with only the duplicated rows
def get_dupes_df(dataframe):
    df = dataframe
    duplicates_df = df[df.duplicated()]

    # Can also pass in specific cols to check for dupes
    # duplicates_df = df[df/duplicated(cols)]

    duplicates_df.to_csv('dupes.csv')